'''
Created on 2016-12-20

@author: sy
'''

#!/usr/bin/python
# -*- coding: UTF-8 -*-
from datetime import date
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis

wb = load_workbook(filename = 'C:\Users\sy\Desktop\statistic\\new2\\graphs.xlsx')
#b = Workbook()
ws = wb.get_sheet_by_name('newsum')
'''
rows = [
    ['Date', 'Batch 1', 'Batch 2', 'Batch 3'],
    [date(2015,9, 1), 40, 30, 25],
    [date(2015,9, 2), 40, 25, 30],
    [date(2015,9, 3), 50, 30, 45],
    [date(2015,9, 4), 30, 25, 40],
    [date(2015,9, 5), 25, 35, 30],
    [date(2015,9, 6), 20, 40, 35],
]

for row in rows:
    ws.append(row)
'''
c1 = LineChart()
c1.title = "Activity覆盖率随时间变化关系"
c1.style = 13
c1.y_axis.title = '覆盖率'
c1.x_axis.title = '时间/分钟'

data = Reference(ws, min_col=1, min_row=2, max_col=61, max_row=5)
c1.add_data(data, from_rows=True,titles_from_data=True)

# Style the lines
'''
s1 = c1.series[0]
s1.marker.symbol = "triangle"
s1.marker.graphicalProperties.solidFill = "FF0000" # Marker filling
s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline
s1.graphicalProperties.line.width = 15005 # width in EMUs

s1.graphicalProperties.line.noFill = True
'''

s1 = c1.series[0]
s1.smooth = True # Make the line smooth
s1.graphicalProperties.line.width = 25005 # width in EMUs

s2 = c1.series[1]
s2.graphicalProperties.line.solidFill = "0000AA"
s2.graphicalProperties.line.dashStyle = "sysDot"
s2.graphicalProperties.line.width = 25005 # width in EMUs

s3 = c1.series[2]
s3.graphicalProperties.line.solidFill = "00AA00"
s3.graphicalProperties.line.dashStyle = "dashDot"
s3.graphicalProperties.line.width = 25005 # width in EMUs

s4 = c1.series[3]
s4.graphicalProperties.line.solidFill = "AA0000"
s4.graphicalProperties.line.dashStyle = "sysDashDot"
s4.graphicalProperties.line.width = 25005 # width in EMUs

ws.add_chart(c1, "A20")
'''
from copy import deepcopy
stacked = deepcopy(c1)
stacked.grouping = "stacked"
stacked.title = "Stacked Line Chart"
ws.add_chart(stacked, "A27")

percent_stacked = deepcopy(c1)
percent_stacked.grouping = "percentStacked"
percent_stacked.title = "Percent Stacked Line Chart"
ws.add_chart(percent_stacked, "A44")

# Chart with date axis
c2 = LineChart()
c2.title = "Date Axis"
c2.style = 12
c2.y_axis.title = "Size"
c2.y_axis.crossAx = 500
c2.x_axis = DateAxis(crossAx=100)
c2.x_axis.number_format = 'd-mmm'
c2.x_axis.majorTimeUnit = "days"
c2.x_axis.title = "Date"

c2.add_data(data, titles_from_data=True)
dates = Reference(ws, min_col=1, min_row=2, max_row=7)
c2.set_categories(dates)

ws.add_chart(c2, "A61")
'''
wb.save("C:\Users\sy\Desktop\statistic\\new2\\graphs.xlsx")