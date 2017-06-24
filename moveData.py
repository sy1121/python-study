'''
Created on 2017-5-19

@author: sy
'''
#!/usr/bin/python
# -*- coding: UTF-8 -*-
import os
import xlrd  # @UnresolvedImport
from openpyxl import Workbook  # @UnresolvedImport
from openpyxl import load_workbook
def getData(filepath):
    f=open(filepath, 'r')
    dict={}
    for  eachLine in f:
        datalist=eachLine.split("|")
        dict[datalist[0].strip()]=datalist[1:9]
    return dict

if __name__ == '__main__':
   filePathC = "C:\Users\sy\Desktop\\adafa.txt"
   dict=getData(filePathC) 
   workbook= load_workbook(filename='C:\Users\sy\Desktop\\coverage.xlsx') 
   sheet = workbook.get_sheet_by_name('Sheet1')
   for key in dict:
       print "%s: %s" % (key,dict[key])
       for index in range(3,59):
           if key ==  str(sheet.cell(row=index,column=1).value).strip():
                sheet.cell(row=index,column=2).value = int(str(dict[key][0]).strip())
                sheet.cell(row=index,column=3).value = int(str(dict[key][1]).strip())
                sheet.cell(row=index,column=4).value = int(str(dict[key][2]).strip())
                sheet.cell(row=index,column=5).value = int(str(dict[key][3]).strip())
                sheet.cell(row=index,column=7).value = int(str(dict[key][4]).strip())
                sheet.cell(row=index,column=8).value = int(str(dict[key][5]).strip())
                sheet.cell(row=index,column=9).value = int(str(dict[key][6]).strip())
                sheet.cell(row=index,column=10).value = int(str(dict[key][7]).strip())
                break       
   workbook.save("C:\Users\sy\Desktop\\coverage.xlsx")
