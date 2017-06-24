'''
Created on 2016-12-19

@author: sy
'''
#!/usr/bin/python
# -*- coding: UTF-8 -*-
import os
import xlrd  # @UnresolvedImport
from openpyxl import Workbook  # @UnresolvedImport

def getActivityInfo(filepath,wb):
    pathDir =  os.listdir(filepath)
    dict={}
    for allDir in pathDir:
        child = os.path.join('%s\%s' % (filepath, allDir))
        print child.decode('gbk') #
        list=[]
        if os.path.isdir(child):
            list=readFile('%s\%s' % (child, "actr.txt")) 
            sum=getSumAct('%s\%s' % (child, "acts.txt"))
            list.insert(0, sum)
        dict[allDir]=list
    writeToExcel(dict,"Activity",wb)
    return dict

def getMethodInfo(filepath,wb):
    pathDir =  os.listdir(filepath)
    dict={}
    for allDir in pathDir:
        child = os.path.join('%s\%s' % (filepath, allDir))
        print child.decode('gbk') #
        list=[]
        if os.path.isdir(child):
            list=readFile('%s\%s' % (child, "methr.txt")) 
            sum=0 ##getSumAct('%s\%s' % (child, "acts.txt"))
            list.insert(0, sum)
        dict[allDir]=list
    writeToExcel(dict,"Method",wb)
    return dict

def readFile(filename):
    try:
        list=[]
        fopen = open(filename, 'r') 
        for eachLine in fopen:
            if ("count" not in eachLine):
                continue
            else:
                if eachLine.partition(':')[2]=="":
                    continue
                list.append(int(eachLine.partition(':')[2]))
    except IOError,ValueError:
        print "Error: 没有找到文件或读取文件失败"
   # finally:
   #     fopen.close()
    return list

def getSumAct(filename):
    try:
        print filename
        actsum=0
        f=open(filename, 'r')
        for  eachLine in f:
            if "count" in eachLine:
                actsum=int(eachLine.partition(':')[2])
                break
    except IOError,ValueError:
        print "Error: 没有找到文件或读取文件失败"
  #  finally:
  #      f.close()
    return actsum
         
def writeFile(filename):
    fopen = open(filename, 'w')
    print "\rplease input mul"," s"
    while True:
        aLine = raw_input()
        if aLine != ".":
            fopen.write('%s%s' % (aLine, os.linesep))
        else:
            print "save file!"
            break
    fopen.close()
    
def writeToExcel(dict,sheetName,wb):
    # grab the active worksheet
    ws = wb.create_sheet(sheetName)
    ws['A1'] = "应用名称"
    ws['B1'] = "sum"
    for index in range(3,63):
        ws.cell(row=1,column=index).value=index-2
    i=2;
    for key in dict:
        ws.cell(row=i,column=1).value=key
        for value in range(2,len(dict[key])+2):
            ws.cell(row=i,column=value).value=dict[key][value-2]
        i=i+1;
    return 

if __name__ == '__main__':
    filePathC = "D:\FDroid-38\\manual-output\\business"#D:\FDroid-38\H1\output"
    wb = Workbook()
    getActivityInfo(filePathC,wb)
    getMethodInfo(filePathC,wb)
    # Save the file
    wb.save("C:\Users\sy\Desktop\statistic\\new2\\manual-output-7.xlsx")
