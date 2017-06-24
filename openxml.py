'''
Created on 2016-12-19

@author: sy
'''
#!/usr/bin/python
# -*- coding: UTF-8 -*-
from openpyxl import Workbook  # @UnresolvedImport
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
for index in range(1,61):
    ws.cell(row=4, column=index).value=index
    ws.cell(row=0,column=index).value=index
# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")