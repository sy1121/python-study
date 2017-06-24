'''
Created on 2017-5-19

@author: sy
'''
#!/usr/bin/python
# -*- coding: UTF-8 -*-
import os
import xlrd  # @UnresolvedImport
from openpyxl import Workbook  # @UnresolvedImport
from openpyxl import load_workbook
if __name__ == '__main__':
   workbook1= load_workbook(filename='C:\Users\sy\Desktop\statistic\\new2\\result2.xlsx') 
   sheet1 = workbook1.get_sheet_by_name('table2')
   workbook2= load_workbook(filename='C:\Users\sy\Desktop\\coverage.xlsx') 
   sheet2 = workbook2.get_sheet_by_name('Sheet1')
   for index1 in range(3,59):
       key=sheet2.cell(row=index1,column=1).value
       print "key: ", key
       for index2 in range(3,61):
           print "cell: ",str(sheet1.cell(row=index2,column=1).value).strip(),str(sheet1.cell(row=index2,column=1).value).strip().lower().startswith(str(key).strip().lower())
           if str(sheet1.cell(row=index2,column=1).value).strip().lower().startswith(str(key).strip().lower()):
                 sheet2.cell(row=index1,column=12).value = round(float(sheet1.cell(row=index2,column=12).value)*100)
                 sheet2.cell(row=index1,column=13).value = round(float(sheet1.cell(row=index2,column=13).value)*100)
                 print "find"
                 break
   workbook2.save("C:\Users\sy\Desktop\\coverage.xlsx")