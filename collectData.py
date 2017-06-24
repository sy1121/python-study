'''
Created on 2016-12-19

@author: sy
'''
#!/usr/bin/python
# -*- coding: UTF-8 -*-
import os
import xlrd  # @UnresolvedImport
from openpyxl import Workbook  # @UnresolvedImport
from openpyxl import load_workbook
def getSum(fileName):
     methsum=0;
     f=open(fileName, 'r')
     for  eachLine in f:
        if "total" in eachLine:
            methsum=eachLine.split(" ")[2]  #.partition(' ')[2].partition(' ')[2]
            break
     return methsum

def getData(filepath):
    pathDir =  os.listdir(filepath)
    dict={}
    for allDir in pathDir:
        child = os.path.join('%s\%s' % (filePathC, allDir))
        #print child.decode('gbk') #
        if os.path.isfile(child):
            sum=getSum(child) 
            dict[allDir]=sum
    return dict

if __name__ == '__main__':
   filePathC = "D:\FDroid-38\covids"
   outPath="D:\FDroid-38\meths.txt"
   dict=getData(filePathC) 
   fopen = open(outPath, 'w')
  # workbook = load_workbook(filename="C:\Users\sy\Desktop\statistic\\new1\HFD3.xlsx")
   workbook= load_workbook(filename='C:\Users\sy\Desktop\statistic\\new2\\manual-output-20.xlsx') 
   sheet = workbook.get_sheet_by_name('Method')
   for key in dict:
       print "%s: %s" % (key,dict[key])
       fopen.write("%s: %s" % (key,dict[key]))
       for index in range(2,47):
               newkey = key.split('_')[0]
               print "%s: %s :%s" % (newkey,sheet.cell(row=index,column=1).value,dict[key])
               if newkey ==  sheet.cell(row=index,column=1).value:
                    sheet.cell(row=index,column=2).value = int(dict[key])
                    break
   workbook.save("C:\Users\sy\Desktop\statistic\\new2\\manual-output-20.xlsx")
   fopen.close()
    