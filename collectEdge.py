'''
Created on 2016-12-19

@author: sy
'''
#!/usr/bin/python
# -*- coding: UTF-8 -*-
import os
import xlrd  # @UnresolvedImport
from openpyxl import Workbook  # @UnresolvedImport
from openpyxl import load_workbook
def getSum(fileName):
     methsum=0;
     f=open(fileName, 'r')
     for  eachLine in f:
        if eachLine.startswith("//e"):
            methsum=methsum+1  #.partition(' ')[2].partition(' ')[2]
     return methsum

def getData(filepath):
    pathDir =  os.listdir(filepath)
    dict={}
    txtFile="edge2event.txt"
    for allDir in pathDir:
        child = os.path.join('%s\%s\%s' % (filePathC, allDir,txtFile))
        #print child.decode('gbk') #
        if os.path.isfile(child):
            sum=getSum(child) 
            dict[allDir]=sum
    return dict

if __name__ == '__main__':
   filePathC = "D:\FDroid-38\\manual-output\\business"
   outPath="D:\FDroid-38\hstates.txt"
   dict=getData(filePathC) 
   fopen = open(outPath, 'w')
  # workbook = load_workbook(filename="C:\Users\sy\Desktop\statistic\\new1\HFD3.xlsx")
   workbook= load_workbook(filename='C:\Users\sy\Desktop\statistic\\new2\\manual-output-7.xlsx') 
   sheet = workbook.get_sheet_by_name('Activity')
   for key in dict:
       print "%s: %s" % (key,dict[key])
       fopen.write("%s: %s" % (key,dict[key]))
       for index in range(2,8):
               print "%s: %s :%s" % (key,sheet.cell(row=index,column=1).value,dict[key])
               if key ==  sheet.cell(row=index,column=1).value:
                    sheet.cell(row=index,column=4).value = int(dict[key])
                    break
   workbook.save("C:\Users\sy\Desktop\statistic\\new2\\manual-output-7.xlsx")
   fopen.close()
    