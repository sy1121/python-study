'''
Created on 2017-5-9

@author: sy
'''
#!/usr/bin/python
# -*- coding: UTF-8 -*-
import os
import xlrd  # @UnresolvedImport
from openpyxl import Workbook  # @UnresolvedImport
from openpyxl import load_workbook
if __name__ == '__main__':
   workbook1= load_workbook(filename='C:\Users\sy\Desktop\statistic\\new2\\result-z.xlsx') 
   sheet1 = workbook1.get_sheet_by_name('table2')
   workbook2= load_workbook(filename='C:\Users\sy\Desktop\statistic\\new2\\11.xlsx') 
   sheet2 = workbook2.get_sheet_by_name('Sheet3')
   for index1 in range(2,59):
       key=sheet2.cell(row=index1,column=1).value
       for index2 in range(2,61):
           if key ==  sheet1.cell(row=index2,column=1).value:
                 sheet2.cell(row=index1,column=4).value = sheet1.cell(row=index2,column=4).value
                 break
   workbook2.save("C:\Users\sy\Desktop\statistic\\new2\\11.xlsx")